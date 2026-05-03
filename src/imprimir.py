from openpyxl import Workbook
from openpyxl.styles import Border,Side,Font
from openpyxl.utils import get_column_letter

from config import data_formatada
from config import categorias
from config import valor_total

#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#FUNÇÃO DE IMPRIMIR NO EXCEL

def escrever_bloco(ws, coluna, linha_inicio, titulo, lista):
    ws.cell(row=linha_inicio, column=coluna, value=titulo).font = Font(bold=True)

    linha = linha_inicio + 1

    for item in lista:
        ws.cell(row=linha, column=coluna, value=item)
        linha += 1

    return linha



def imprimir():
    wb = Workbook()
    ws = wb.active
    ws.title = "Intenções"

    # Título principal
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    titulo = ws["A1"]
    titulo.value = f"INTENÇÕES DA MISSA {data_formatada}"
    titulo.font = Font(bold=True)
    

    layout = [
        (1, 2, "HONRA E LOUVOR"),
        (1, None, "ANIVERSÁRIO NATALÍCIO"),
        (1, None, "ANIVERSÁRIO DE CASAMENTO"),
        (1, None, "AÇÃO DE GRAÇAS"),

        (2, 2, "INTENÇÃO"),
        (2, None, "RECUPERAÇÃO DE SAÚDE"),

        (3, 2, "7º DIA"),
    ]

    proxima_linha = {}

    # Blocos normais
    for coluna, linha_inicio, titulo in layout:

        if linha_inicio is not None:
            linha = linha_inicio
        else:
            linha = proxima_linha[coluna]

        lista = categorias.get(titulo, [])
        proxima_linha[coluna] = escrever_bloco(ws, coluna, linha, titulo, lista) + 1

    # IRMÃOS FALECIDOS (dividido em duas colunas)
    falecidos = categorias.get("IRMÃOS FALECIDOS", [])

    metade = (len(falecidos) + 1) // 2
    lista_col3 = falecidos[:metade]
    lista_col4 = falecidos[metade:]

    linha_base = 8

    fim_col3 = escrever_bloco(ws, 3, linha_base, "IRMÃOS FALECIDOS", lista_col3)
    fim_col4 = escrever_bloco(ws, 4, linha_base, "IRMÃOS FALECIDOS", lista_col4)

    max_linha = max(ws.max_row, fim_col3, fim_col4)

    # Bordas (até coluna 4, mesmo vazias)
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in range(1, max_linha + 1):
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = borda

    # Largura ideal para A4
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 38

    wb.save("output/Intenções.xlsx")

    print("\nArquivo Excel 'Intenções.xlsx' criado com sucesso na pasta output!\n")


#-------------------------------------------------------------------------------------------------------------

    
    sair_menu2 = 1

