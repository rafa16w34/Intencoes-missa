from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Border,Side,Font
from openpyxl.utils import get_column_letter
from docx import Document

data_hoje  = datetime.now()

data_ajustada = data_hoje + timedelta(hours=1)#Soma 1 na hora porque temos que imprimir 10 minutos antes
data_formatada = data_ajustada.strftime("%d/%m/%Y - %H:00")

#Listas para cada uma das intenções
categorias = {
    "HONRA E LOUVOR": [],
    "INTENÇÃO": [],
    "ANIVERSÁRIO NATALÍCIO": [],
    "RECUPERAÇÃO DE SAÚDE": [],
    "ANIVERSÁRIO DE CASAMENTO": [],
    "AÇÃO DE GRAÇAS": [],
    "IRMÃOS FALECIDOS": [],
    "7º DIA": []
}

pix = []

#Variavel do dinheiro
valor_total = 0

def quebrar_linha(texto, limite=35):
    palavras = texto.split()
    linhas = []
    linha_atual = ""

    for palavra in palavras:
        # testa se a palavra cabe na linha atual
        if len(linha_atual) + len(palavra) + 1 <= limite:
            if linha_atual:
                linha_atual += " "
            linha_atual += palavra
        else:
            # fecha a linha atual com "-"
            linhas.append(linha_atual + "-")
            linha_atual = "-" + palavra

    # adiciona a última linha
    if linha_atual:
        linhas.append(linha_atual)

    return linhas

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#FUNÇÃO DE ESCANEAR O DOCUMENTO WORD

def extrair_docx():

    def normalizar(txt):
        return " ".join(txt.strip().upper().split())

    arquivo = "intencoes.docx"
    doc = Document(arquivo)

    TITULOS = [
        "HONRA E LOUVOR",
        "INTENÇÃO",
        "ANIVERSÁRIO NATALÍCIO",
        "ANIVERSÁRIO DE CASAMENTO",
        "RECUPERAÇÃO DE SAÚDE",
        "AÇÃO DE GRAÇAS",
        "IRMÃOS FALECIDOS",
    ]

    TITULOS_NORM = [normalizar(t) for t in TITULOS]

    dados = {}
    titulo_atual = None

    # LER TABELAS EM VEZ DE PARÁGRAFOS
    for table in doc.tables:
        for row in table.rows:
            for cell in row.cells:

                linha = cell.text.strip()
                if not linha:
                    continue

                ln = normalizar(linha)

                if ln in TITULOS_NORM:  # detecta título
                    titulo_atual = ln
                    dados[titulo_atual] = []
                    continue

                if titulo_atual:
                    
                    if (linha != "+") and not (linha.startswith("INTENÇÕES")):

                        linhas_quebradas = quebrar_linha(linha)

                        for l in linhas_quebradas:
                            dados[titulo_atual].append(l)



    return dados



dados_importados = extrair_docx()



for titulo, itens in dados_importados.items():

    for item in itens:

        # se NÃO começar com 7º, mantém na categoria original
        if not item.startswith("7º"):
            categorias[titulo].append(item)

        # se começar com 7º, manda para a categoria 7º DIA
        else:
            categorias["7º DIA"].append(item)


#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

def adicionar_intencao(categoria, texto):
    linhas = quebrar_linha(texto.title())
    for linha in linhas:
        categorias[categoria].append(linha)


#FUNÇÃO DE MARCAR MISSA

def menu_pessoa(valor_pessoa):

    sair_menu2 = 0

    categoria = None
    texto = None

    while sair_menu2 == 0:#Menu para marcar a intenção de uma pessoa

        try:

            opcao = int(input('\nDigite uma das opções abaixo:\n1- Honra e Louvor\n2- Aniversário\n3- Aniversário de Casamento\n4- Ação de Graças\n5- Intenções\n6- Recuperação de Saúde\n7- 7º Dia\n8- Almas\n------------------------------------------------------------\n9- Voltar\n-> '))

            match(opcao):

                case 1:#Adiciona uma intenção de Honra e Louvor

                    categoria = "HONRA E LOUVOR"

                    texto = '\nDigite a intenção em Honra e Louvor:\n-> '


                case 2:#Adiciona uma intenção para Aniversário

                    categoria = "ANIVERSÁRIO NATALÍCIO"

                    texto = '\nDigite o nome do aniversariante:\n-> '


                case 3:#Adiciona uma intenção de Aniversário de Casamento
                    
                    categoria = "ANIVERSÁRIO DE CASAMENTO"

                    texto = '\nDigite o nome do casal aniversariante:\n-> '



                case 4:#Adiciona uma intenção de Ação de Graças
                    
                    categoria = "AÇÃO DE GRAÇAS"
                            
                    texto = '\nDigite a intenção de acão de graças:\n-> '



                case 5:#Adiciona uma intenção de Intenções
                    
                    categoria = "INTENÇÃO"
                            
                    texto = '\nDigite a intenção:\n-> '




                case 6:#Adiciona uma intenção de Recuperação de Saúde
                    
                    categoria = "RECUPERAÇÃO DE SAÚDE"

                    texto = '\nDigite o nome do enfermo:\n-> '

                    


                case 7:#Adiciona uma intenção de Sétimo dia
                    
                    categoria = "7º DIA"

                    texto = '\nDigite o nome do sétimo dia:\n-> '

                   

                case 8:#Adiciona uma intenção de Alma
                    
                    categoria = "IRMÃOS FALECIDOS"

                    texto = '\nDigite o nome da alma:\n-> '

                    

                case 9:

                
                    print(f'\nO valor a ser pago é de R$ {valor_pessoa:.2f}\n')

                    metodo_de_pagamento = int(input('\nQual vai ser a forma de pagamento?\n1- Dinheiro\n2- Pix\n-> '))

                    match(metodo_de_pagamento):

                        case 1:#Dinheiro

                            while True:
                                
                                try:

                                    valor_recebido = int(input('\nDigite o valor recebido:\n-> R$ '))

                                    if (valor_recebido < valor_pessoa):

                                        print('\nDigite um valor válido!\n')


                                    else:
                                        
                                        if (valor_recebido > valor_pessoa):

                                            print(f'\nTroco: R$ {(valor_recebido - valor_pessoa):.2f}\n')
                                        
                                        break

                                
                                
                                except(ValueError):
                                    
                                    print('\nDigite um valor válido!\n')

                        case 2:#Pix

                            nome_pix = str(input("\nQual é o nome quem irá fazer o pix?\n-> "))

                            nome_pix = nome_pix.lower()

                            pix.append(nome_pix.title())
                            pix.append(valor_pessoa)

                            with open ('Lista_Pix.txt','w') as arq:

                                arq.write(f'Lista do Pix {data_formatada}:\n')

                                for i in range(len(pix)):

                                    if ((i+1)%2 != 0):                                    
                                        arq.write(f'- {pix[i]} : R$ {pix[i+1]},00\n')
                                        i+2
                                    
                                    if i+2 >= len(pix):
                                        break



                    confirmacao = str(input('\nAperte enter para continuar\n'))

                    print('\nVoltando ao menu inicial...\n')

                    sair_menu2 = 1
                

                case _: 

                    print('\nErro: Digite uma das opções mostradas no menu!\n')
        

        except(ValueError):
            print('\nErro: Digite uma das opções mostradas no menu!\n')
            continue


        if opcao != 9:

            nome = str(input(texto))
            nome = nome.lower()

            adicionar_intencao(categoria, nome)
            valor_pessoa += 3

    
    return(valor_pessoa)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

#FUNÇÃO DE IMPRIMIR NO EXCEL

def escrever_bloco(ws, coluna, linha_inicio, titulo, lista):
    ws.cell(row=linha_inicio, column=coluna, value=titulo).font = Font(bold=True)

    linha = linha_inicio + 1

    for item in lista:
        ws.cell(row=linha, column=coluna, value=item)
        linha += 1

    return linha



def imprimir():
    wb = Workbook()
    ws = wb.active
    ws.title = "Intenções"

    # Título principal
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    titulo = ws["A1"]
    titulo.value = f"INTENÇÕES DA MISSA {data_formatada}"
    titulo.font = Font(bold=True)
    

    layout = [
        (1, 2, "HONRA E LOUVOR"),
        (1, None, "ANIVERSÁRIO NATALÍCIO"),
        (1, None, "ANIVERSÁRIO DE CASAMENTO"),
        (1, None, "AÇÃO DE GRAÇAS"),

        (2, 2, "INTENÇÃO"),
        (2, None, "RECUPERAÇÃO DE SAÚDE"),

        (3, 2, "7º DIA"),
    ]

    proxima_linha = {}

    # Blocos normais
    for coluna, linha_inicio, titulo in layout:

        if linha_inicio is not None:
            linha = linha_inicio
        else:
            linha = proxima_linha[coluna]

        lista = categorias.get(titulo, [])
        proxima_linha[coluna] = escrever_bloco(ws, coluna, linha, titulo, lista) + 1

    # IRMÃOS FALECIDOS (dividido em duas colunas)
    falecidos = categorias.get("IRMÃOS FALECIDOS", [])

    metade = (len(falecidos) + 1) // 2
    lista_col3 = falecidos[:metade]
    lista_col4 = falecidos[metade:]

    linha_base = 8

    fim_col3 = escrever_bloco(ws, 3, linha_base, "IRMÃOS FALECIDOS", lista_col3)
    fim_col4 = escrever_bloco(ws, 4, linha_base, "IRMÃOS FALECIDOS", lista_col4)

    max_linha = max(ws.max_row, fim_col3, fim_col4)

    # Bordas (até coluna 4, mesmo vazias)
    borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in range(1, max_linha + 1):
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = borda

    # Largura ideal para A4
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 38

    wb.save("Intenções.xlsx")

    print("\nArquivo Excel 'Intenções.xlsx' criado com sucesso!\n")


#-------------------------------------------------------------------------------------------------------------
    print(f'\nO valor total do dia é de: R$ {valor_total:.2f}\n')
    
    sair_menu2 = 1

#-------------------------------------------------------------------------------------------------------------

#FUNÇÃO DE EDITAR A INTENÇÃO JÁ MARCADA

def editar(lista):

    divida = 0

    if (len(lista) != 0):
        
        for i in range(len(lista)+1):

            if (i < len(lista)):
                print(f'\n{i} : {lista[i]}') #Lista todas as intenções marcadas no tipo escolhido com o indice na frente
            else:
                print(f'\n{i} : Sair')
            
        print('\nEscolha qual será editado pelo índice:\n')

        while True:

            try:
                edicao = int(input('-> '))

            except ValueError:
                print('\nDigite um valor válido!\n')
                continue

            # Verifica se o índice existe
            if (0 <= edicao) and (edicao < len(lista)):
                sair = False
                break

            elif (edicao >= len(lista)):
                sair = True
                break

            else:

                print('\nÍndice inválido! Tente novamente.\n')

        while sair == False:

            print('\nGostaria de remover, editar ou sair?\n')
            opcao_edicao = input('-> ').lower()

            if opcao_edicao == 'remover':
                lista.pop(edicao)
                divida -= 3
                print("\nIntenção removida com sucesso!\n")
                break

            elif opcao_edicao == 'editar':
                nova_intencao = input('\nDigite a nova intenção:\n-> ')
                lista[edicao] = nova_intencao
                print('\nIntenção editada com sucesso!\n')
                break

            elif opcao_edicao == 'sair':
                break

            else:
                print('Digite "remover" ou "editar" !')

    else:
        print('\nErro: Lista vazia!\n')

    return(divida)


#-------------------------------------------------------------------------------------------------------------

#MENU PRINCIPAL

while True:
    try:
        opcao = int(input(
            '\nDigite uma das opções a seguir:\n'
            '1- Marcar intenção\n'
            '2- Dados\n'
            '3- Editar\n'
            '4- Imprimir\n'
            '0- Finalizar programa\n-> '
        ))
    except ValueError:
        print('\nErro: Digite um número válido!\n')
        continue

    match opcao:

        case 1:

            valor_pessoa = 0
            

            valor_total += menu_pessoa(valor_pessoa)

        case 2:

            print(f'\nDados atuais:\nTotal de intenções: {valor_total/3}\nValor total: R$ {valor_total:.2f}')

        case 3:

            try:
                opcao_editar = int(input('\nGostaria de editar qual dos tipos de intenção?\n1- Honra e Louvor\n2- Aniversário\n3- Aniversário de Casamento\n4- Ação de Graças\n5- Intenções\n6- Recuperação de Saúde\n7- 7º Dia\n8- Almas\n9- Voltar\n-> '))
                
                lista = None

                match(opcao_editar):

                    case 1:
                        lista = categorias['HONRA E LOUVOR']

                    case 2:
                        lista = categorias['ANIVERSÁRIO NATALÍCIO']

                    case 3:
                        lista = categorias['ANIVERSÁRIO DE CASAMENTO']

                    case 4:
                        lista = categorias['AÇÃO DE GRAÇAS']

                    case 5:
                        lista = categorias['INTENÇÃO']

                    case 6:
                        lista = categorias['RECUPERAÇÃO DE SAÚDE']

                    case 7:
                        lista = categorias['7º DIA']

                    case 8:
                        lista = categorias['IRMÃOS FALECIDOS']

                    case 9:
                        print('\nVoltando ao menu..\n')
                        continue

                    case _:
                        print('\nDigite uma das opções exibidas no menu!\n')
                        continue

                if lista is not None:
                    valor_total += editar(lista)


            except(ValueError):
                print('\nDigite um valor válido!\n')    


        case 4:#Imprime a lista formatada
            

            imprimir()


        case 0:

            confirmacao = str(input('\nDigite "sim" para sair.\n(Tudo o que foi digitado será perdido caso não tenha sido impresso!)\n-> '))

            if (confirmacao == 'sim'):
                print('\nFinalizando programa...\n')
                break



        case _: 

            print('\nErro: Digite uma das opções mostradas no menu!\n')
